import os
import sys
import openpyxl
import pandas as pd
import argparse

def flag(log):
    # 初期化
    read_flag = 0
    write_flag = 0
    # ファイルを読み込みモードで開く
    with open(log, 'r') as file:
    # ファイルの各行に対して処理
        for line in file:
            # "read :"が行に含まれている場合、read_flagを1に設定
            if "read :" in line:
                read_flag = 1
            # "write :"が行に含まれている場合、write_flagを1に設定
            if "write :" in line:
                write_flag = 1
    return read_flag, write_flag

def search_lines(file_path, targets, read_flag, write_flag):
    found_read_line_numbers = []
    found_write_line_numbers = []

    with open(file_path, 'r') as file:
        lines = file.readlines()

    found_readwrite_line_numbers = []

    for target in targets:
        target_found = False
        for i, line in enumerate(lines, start=1):
            if target in line:
                target_found = True
                found_readwrite_line_numbers.append(i)

        if not target_found:
            print(f"Target '{target}' not found in the file.")

    if read_flag == 1 and write_flag == 1:
        if found_readwrite_line_numbers:
            found_read_line_numbers = found_readwrite_line_numbers[::2]
            found_write_line_numbers = found_readwrite_line_numbers[1::2]
        else:
            found_read_line_numbers = []
            found_write_line_numbers = []
    else:
        found_read_line_numbers = found_readwrite_line_numbers if read_flag == 1 else []
        found_write_line_numbers = found_readwrite_line_numbers if write_flag == 1 else []

    # 要素数を合わせるために0を追加
    max_length = max(len(found_read_line_numbers), len(found_write_line_numbers))
    found_read_line_numbers += [0] * (max_length - len(found_read_line_numbers))
    found_write_line_numbers += [0] * (max_length - len(found_write_line_numbers))

    return found_read_line_numbers, found_write_line_numbers

#指定された行数から文字列抽出
def extract_text_from_file(filename, line_numbers, keywords_to_search):
    result = []

    with open(filename, 'r') as file:
        lines = file.readlines()

    for line_number, keyword in zip(line_numbers, keywords_to_search):
        if 1 <= line_number <= len(lines):
            current_line = lines[line_number - 1].strip()

            if keyword in current_line:
                keyword_index = current_line.find(keyword)
                start_index = keyword_index + len(keyword)
                extracted_text = ''

                for char in current_line[start_index:]:
                    if char.isdigit() or char == '.':
                        extracted_text += char
                    elif char.isspace():
                        continue
                    else:
                        break

                print(f'Debug: 行 {line_number} - {current_line}')
                print(f'Debug: 開始 - 行番号: {line_number}, キーワード: {keyword}')
                print(f'Debug: 結果に追加 - 行番号: {line_number}, 抽出テキスト: {extracted_text}')

                result.append((line_number, extracted_text))
            else:
                print(f'Warning: 行 {line_number} に対応するキーワードがありません。')

        else:
            print(f'Warning: 行 {line_number} は存在しません。')
            result.append((line_number, "0"))

    return result

#nines専用
def find_line_numbers(file_path, target_keywords):
    line_numbers = {keyword: [] for keyword in target_keywords}

    with open(file_path, 'r') as file:
        current_line_number = 0
        for line in file:
            current_line_number += 1
            for keyword in target_keywords:
                if re.search(re.escape(keyword), line):  # 正規表現を使用して部分一致を確認
                    line_numbers[keyword].append(current_line_number)

    # 行番号のみのリストに変換
    line_numbers = {keyword: numbers for keyword, numbers in line_numbers.items() if numbers}

    return line_numbers



def get_log_files_in_directories(directory_path,warikomi,moji,kansu,summary_sheet,wb,targets,keywords_to_search):
    log_files_list = []
    subfolder_count = len([name for name in os.listdir(directory_path) if os.path.isdir(os.path.join(directory_path, name))])

    for i in range(1, subfolder_count + 1):
        current_folder = os.path.join(directory_path, str(i))




        if os.path.exists(current_folder) and os.path.isdir(current_folder):
            #result_*.logがあるなら以下の処理
            log_files = [file for file in os.listdir(current_folder) if file.startswith('result_') and file.endswith('.log')]

            if log_files:
                log = os.path.join(directory_path, str(i), log_files[0])
                read_flag, write_flag = flag(log)
                found_read, found_write = search_lines(log, targets, read_flag, write_flag)
                log_files_list.append(log)
                read_result = extract_text_from_file(log, found_read, keywords_to_search)
                write_result = extract_text_from_file(log, found_write, keywords_to_search)
                read_numbers = [float(t[1]) for t in read_result]
                write_numbers = [float(t[1]) for t in write_result]
                write_excel_summary(summary_sheet,warikomi,moji,kansu,read_numbers)
                write_excel_summary(summary_sheet,warikomi,moji,kansu,write_numbers)
                print(log)
            
            #csvファイルがあるなら以下の処理
            csv_files = [file for file in os.listdir(current_folder) if file.endswith('.csv')]
            if csv_files:
                csv = os.path.join(directory_path, str(i), csv_files[0])
                sheet_name = str(i)
                wb.create_sheet(sheet_name)
                log_files_list.append(csv)
                df = pd.read_csv(csv)
                #sheet = wb[sheet_name]
                #ws = wb.active
                existing_sheet = wb[sheet_name]
                # シート名を指定する場合
                #ws.title = sheet_name
                existing_sheet["G1"] = "=IFERROR(AVERAGEIFS(A:A, D:D, 1), "")"
                existing_sheet["H1"] = "=IFERROR(MAXIFS(A:A, D:D, 1), "")"
                existing_sheet["I1"] = "=IFERROR(MINIFS(A:A, D:D, 1), "")"
                existing_sheet["J1"] = '=LARGE(INDIRECT("A1:A"&COUNTIF(D:D,"=1")),10)'
                existing_sheet["K1"] = "=IFERROR(AVERAGEIFS(A:A, D:D, 0), "")"
                existing_sheet["L1"] = "=IFERROR(MAXIFS(A:A, D:D, 0), "")"
                existing_sheet["M1"] = "=IFERROR(MINIFS(A:A, D:D, 0), "")"
                existing_sheet["N1"] = '=LARGE(INDIRECT("A" & COUNTIF(D:D, "=1") + 1 & ":A" & (COUNTIF(D:D, "=1") + COUNTIF(D:D, "=0"))), 10)'
                # DataFrameのデータをシートに書き込む
                for r_idx, row in enumerate(df.values, 1):
                    for c_idx, value in enumerate(row, 1):
                        existing_sheet.cell(row=r_idx, column=c_idx, value=value)

                print(csv)
    return log_files_list

def main(summary_sheet,wb):
    parser = argparse.ArgumentParser(description='Get paths of .log files in subdirectories.')
    parser.add_argument('directory_path', type=str, help='Path to the directory containing subdirectories.')

    args = parser.parse_args()
    directory_path = args.directory_path
    #変数関係
    targets = ['clat (usec): min=', 'clat (usec): min=', 'clat (usec): min=','clat (usec): min=',' lat (usec): min=',' lat (usec): min=',' lat (usec): min=',' lat (usec): min=','30.00th=[','30.00th=[']
    keywords_to_search = ["stdev=", "min=", "avg=", "max=","stdev=", "min=", "avg=", "max=",'30.00th=[','40.00th=[']
    warikomi = [1 ,3, 4, 6]
    moji = [6]
    kansu="=sum(A1:A3)"
    result = get_log_files_in_directories(directory_path,warikomi,moji,kansu,summary_sheet,wb,targets,keywords_to_search)
    #print(result)



def write_excel_summary(sheet,warikomi,moji,kansu,nums):
    row = 1
    column = sheet.max_row + 1
    count = 0
    # 渡されたデータをエクセルの1行に書き込む
    for num in nums:
        if count in warikomi:
            if count in moji:
                sheet.cell(row=column, column=row, value=kansu)
            else:
                sheet.cell(row=column, column=row, value=None)
            row += 1
        sheet.cell(row=column, column=row, value=num)
        row += 1
        count += 1



if __name__ == '__main__':
    wb = openpyxl.Workbook()
    wb.save('result.xlsx')
    wb = openpyxl.load_workbook("result.xlsx")
    ws = wb.worksheets[0]
    #ws.title = "Sheet1"
    #summary_sheet = ws.title
    #summary_sheet = wb.active
    ws.title = "summary"
    main(ws,wb)
    wb.save('result.xlsx')
