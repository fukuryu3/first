import re
import openpyxl
# ファイルのパス
file_path = 'result.log'

def flag():
    # 初期化
    read_flag = 0
    write_flag = 0
    # ファイルを読み込みモードで開く
    with open(file_path, 'r') as file:
    # ファイルの各行に対して処理
        for line in file:
            # "read :"が行に含まれている場合、read_flagを1に設定
            if "read :" in line:
                read_flag = 1
            # "write :"が行に含まれている場合、write_flagを1に設定
            if "write :" in line:
                write_flag = 1
    return read_flag, write_flag

def search_lines(file_path, targets, read_flag, write_flag):
    found_read_line_numbers = []
    found_write_line_numbers = []

    with open(file_path, 'r') as file:
        lines = file.readlines()

    found_readwrite_line_numbers = []

    for target in targets:
        target_found = False
        for i, line in enumerate(lines, start=1):
            if target in line:
                target_found = True
                found_readwrite_line_numbers.append(i)

        if not target_found:
            print(f"Target '{target}' not found in the file.")

    if read_flag == 1 and write_flag == 1:
        if found_readwrite_line_numbers:
            found_read_line_numbers = found_readwrite_line_numbers[::2]
            found_write_line_numbers = found_readwrite_line_numbers[1::2]
        else:
            found_read_line_numbers = []
            found_write_line_numbers = []
    else:
        found_read_line_numbers = found_readwrite_line_numbers if read_flag == 1 else []
        found_write_line_numbers = found_readwrite_line_numbers if write_flag == 1 else []

    # 要素数を合わせるために0を追加
    max_length = max(len(found_read_line_numbers), len(found_write_line_numbers))
    found_read_line_numbers += [0] * (max_length - len(found_read_line_numbers))
    found_write_line_numbers += [0] * (max_length - len(found_write_line_numbers))

    return found_read_line_numbers, found_write_line_numbers

#指定された行数から文字列抽出
def extract_text_from_file(filename, line_numbers, keywords_to_search):
    result = []

    with open(filename, 'r') as file:
        lines = file.readlines()

    for line_number, keyword in zip(line_numbers, keywords_to_search):
        if 1 <= line_number <= len(lines):
            current_line = lines[line_number - 1].strip()

            if keyword in current_line:
                keyword_index = current_line.find(keyword)
                start_index = keyword_index + len(keyword)
                extracted_text = ''

                for char in current_line[start_index:]:
                    if char.isdigit() or char == '.':
                        extracted_text += char
                    elif char.isspace():
                        continue
                    else:
                        break

                print(f'Debug: 行 {line_number} - {current_line}')
                print(f'Debug: 開始 - 行番号: {line_number}, キーワード: {keyword}')
                print(f'Debug: 結果に追加 - 行番号: {line_number}, 抽出テキスト: {extracted_text}')

                result.append((line_number, extracted_text))
            else:
                print(f'Warning: 行 {line_number} に対応するキーワードがありません。')

        else:
            print(f'Warning: 行 {line_number} は存在しません。')
            result.append((line_number, "0"))

    return result

#nines専用
def find_line_numbers(file_path, target_keywords):
    line_numbers = {keyword: [] for keyword in target_keywords}

    with open(file_path, 'r') as file:
        current_line_number = 0
        for line in file:
            current_line_number += 1
            for keyword in target_keywords:
                if re.search(re.escape(keyword), line):  # 正規表現を使用して部分一致を確認
                    line_numbers[keyword].append(current_line_number)

    # 行番号のみのリストに変換
    line_numbers = {keyword: numbers for keyword, numbers in line_numbers.items() if numbers}

    return line_numbers
#nines専用

    result = []

    with open(file_path, 'r') as file:
        lines = file.readlines()

    for line_number, keyword in zip(line_numbers, keywords_to_search):
        if 1 <= line_number <= len(lines):
            current_line = lines[line_number - 1].strip()

            if keyword in current_line:
                keyword_index = current_line.find(keyword)
                start_index = keyword_index + len(keyword)
                extracted_text = ''

                for char in current_line[start_index:]:
                    if char.isdigit() or char == '.':
                        extracted_text += char
                    else:
                        break

                print(f'Debug: 行 {line_number} - {current_line}')
                print(f'Debug: 開始 - 行番号: {line_number}, キーワード: {keyword}')
                print(f'Debug: 結果に追加 - 行番号: {line_number}, 抽出テキスト: {extracted_text}')

                result.append((line_number, extracted_text))
                
                # もしline_numbersが逆順になっている場合、最後の行番号に達したら関数を終了
                if line_number == line_numbers[-1]:
                    break
            else:
                print(f'Warning: 行 {line_number} に対応するキーワードがありません。')
        else:
            print(f'Warning: 行 {line_number} は存在しません。')

    return result

# nines関係使用例
nines_target_strings = [' lat (usec): min=', 'clat (usec): min=', '99.99th=[', 'total=r=', 'io=']  # 検索する文字列を適切なものに変更してください
found_line_numbers = find_line_numbers(file_path, nines_target_strings)
all_line_numbers = [number for numbers in found_line_numbers.values() for number in numbers]
print(f'すべての行番号: {all_line_numbers}')
line_numbers = all_line_numbers
#keywords_to_search = ["avg=", "max=", "99.99th=[", "total=r=", "io="]
result = extract_text_from_file(file_path, line_numbers, nines_target_strings)

# 出力結果
print(result)

# 使用例
# 結果を出力
read_flag, write_flag = flag()
print("read_flag:", read_flag)
print("write_flag:", write_flag)
# 使用例
targets = ['clat (usec): min=', 'clat (usec): min=', 'clat (usec): min=','clat (usec): min=',' lat (usec): min=',' lat (usec): min=',' lat (usec): min=',' lat (usec): min=','30.00th=[','30.00th=[']


found_read, found_write = search_lines(file_path, targets, read_flag, write_flag)

print("Found read line numbers:", found_read)
print("Found write line numbers:", found_write)
# 使用例
read_line_numbers = found_read
write_line_numbers = found_write
keywords_to_search = ["stdev=", "min=", "avg=", "max=","stdev=", "min=", "avg=", "max=",'30.00th=[','40.00th=[']
read_result = extract_text_from_file(file_path, read_line_numbers, keywords_to_search)
write_result = extract_text_from_file(file_path, write_line_numbers, keywords_to_search)
# 出力結果
print(read_result)
print(write_result)

read_numbers = [float(t[1]) for t in read_result]
write_numbers = [float(t[1]) for t in write_result]
print(read_numbers)
print(write_numbers)

#python
import openpyxl

def create_excel(nums, warikomi, moji, kansu,filename='output.xlsx'):
    # ワークブックを作成
    try:
        # 既存のワークブックを開く
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # ファイルが存在しない場合、新しいワークブックを作成
        workbook = openpyxl.Workbook()

    # アクティブなシートを取得
    sheet = workbook.active

    # データを書き込む行を指定
    row = 1
    column = sheet.max_row + 1
    count = 0
    # 渡されたデータをエクセルの1行に書き込む
    for num in nums:
        if count in warikomi:
            if count in moji:
                sheet.cell(row=column, column=row, value=kansu)
            else:
                sheet.cell(row=column, column=row, value=None)
            row += 1
        sheet.cell(row=column, column=row, value=num)
        row += 1
        count += 1

    # ファイルを保存
    workbook.save(filename)

nums = [375.48, 166.0, 249.06, 41865.0, 375.48, 166.0, 249.25, 41866.0, 195.0, 199.0]
warikomi = [1 ,3, 4, 6]
moji = [6]
kansu="=sum(A1:A3)"

# 使用例
create_excel(nums, warikomi, moji, kansu,'output.xlsx')
create_excel(nums, warikomi, moji, kansu,'output.xlsx')